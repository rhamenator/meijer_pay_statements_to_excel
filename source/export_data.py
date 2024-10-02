import os
import re
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, NamedStyle, GradientFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from file_utils import file_lock_wait
from tqdm import tqdm
from openpyxl.utils.cell import get_column_letter
import dicttoxml # type: ignore
import json
from main_utils import print_message

# Constants
top_level_keys = ["Identification", "Pay Period Info", "Earnings", "Employee Taxes", "Post Tax Deductions", "Pre Tax Deductions", "Employer Paid Benefits", "Subject or Taxable Wages", "Allowances", "Absence Plans", "Payment Information"]
hair_border = Border(left=Side(style='hair'), 
                right=Side(style='hair'), 
                top=Side(style='hair'), 
                bottom=Side(style='hair'))
fills = [
    PatternFill(start_color="6376C7", end_color="6376C7", fill_type="solid"),  
    PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid"),  # Light Cyan
    PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Light Green
    PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")   # Light Yellow
]
fonts = [
    Font(size=12, bold=True),
    Font(size=11, bold=True),
    Font(size=10, bold=True),
    Font(size=9, bold=True)
]
accounting_format = NamedStyle(name='accounting_format', number_format='_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)')
date_format = NamedStyle(name='date_format', number_format='mm/dd/yyyy')
number_format = NamedStyle(name='number_format', number_format='#,##0.00')
dark_row_fill_color = '98E3E2'
light_row_fill_color = 'C5E3E2'
light_shade = PatternFill(start_color=light_row_fill_color, end_color=light_row_fill_color, fill_type="solid")
medium_shade = PatternFill(start_color=dark_row_fill_color, end_color=dark_row_fill_color, fill_type="solid")

def save_to_excel(data, output_path):
    # global gui_mode, progress_bar, progress
    if not file_lock_wait(output_path):
        return False
    
    print_message(f'Now creating {output_path}')
    num_pages = len(data)
    flattened_data = []
    for idx, page in enumerate(data):
        flat_page = flatten_dict(data)
        flat_page['UniqueID'] = idx + 1
        flattened_data.append(flat_page)

    df = pd.DataFrame(flattened_data)

    try:
        wb = Workbook()
        wb.add_named_style(accounting_format)
        wb.add_named_style(date_format)
        wb.add_named_style(number_format)

        def is_numeric_string(s):
            return bool(re.match(r'^\d+$', s))

        for section in top_level_keys:
            if section == 'General':
                continue

            section_df = df[[col for col in df.columns if col.startswith(section + ':') or col.endswith("Date") or col.endswith("Begin") or col.endswith("End")]]
            if section_df.empty:
                continue
            
            ws = wb.create_sheet(title=section)
            
            max_levels = create_hierarchical_headers(ws, section_df, fills, fonts)
            outfile = os.path.basename(output_path)
            append_rows(section, section_df, ws)
            total_rows = ws.max_row
            sheet_rows = ws.iter_rows(min_row=max_levels + 1, max_row=total_rows, min_col=1, max_col=ws.max_column)
            format_cells(is_numeric_string, section, section_df, max_levels, total_rows, sheet_rows)
            
            adjust_column_widths(section, ws)
            
            format_rows(ws, max_levels, section)

        wb.remove(wb['Sheet'])
        
        if not file_lock_wait(output_path):
            return False
        
        print_message(f'Now saving {os.path.basename(output_path)}')
        wb.save(output_path)
        print_message(f"{os.path.basename(output_path)} has been created.")
                    
    except Exception as e:
        print_message(f"Error creating {os.path.basename(output_path)}: {e}")
        raise

def format_cells(is_numeric_string, section, section_df, max_levels, total_rows, sheet_rows):
    for row in tqdm(sheet_rows, f'Formatting cells in the {section} section', total_rows):
        for cell in row:
            col_name = section_df.columns[cell.column - 1]
            if cell.row <= max_levels:
                cell.alignment = Alignment(horizontal='center')
            else:
                if col_name.endswith(('Rate', 'Amount', 'YTD', 'Pay', 'Deductions', 'Taxes')):
                    cell.style = accounting_format
                elif 'date' in col_name.lower() or 'begin' in col_name.lower() or 'start' in col_name.lower() or 'end' in col_name.lower():
                    cell.style = date_format
                elif isinstance(cell.value, float):
                    cell.style = number_format
                elif isinstance(cell.value, int):
                    cell.number_format = '0'
                elif isinstance(cell.value, str) and is_numeric_string(cell.value):
                    cell.number_format = '0'

def append_rows(section, section_df, ws):
    rows = dataframe_to_rows(section_df, index=False, header=False)
    for r in tqdm(rows, f'Appending data to the {section} section', rows.__sizeof__()):
        ws.append(r)

def format_rows(ws, max_levels, section):
    for row in tqdm(range(max_levels + 1, ws.max_row + 1), f'Formatting rows in the {section} section', ws.max_row + 1):
        fill = light_shade if (row - max_levels) % 2 == 0 else medium_shade
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill

def adjust_column_widths(section, ws):
    columns = ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column)
    for col in tqdm(columns,f'Adjusting column widths in the {section}', ws.max_column):
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width
    
    
def save_to_json(data, output_path):
    if not file_lock_wait(output_path):
        return False
    
    print_message(f'Now creating {os.path.basename(output_path)}')
    try:
        with open(output_path, 'w', encoding='utf-8') as json_file:
            json.dump(data, json_file, ensure_ascii=False, indent=4)
        print_message(f'Output saved to JSON: {os.path.basename(output_path)}')
        return True
    except Exception as e:
        print_message(f"Error saving to JSON: {e}")
        raise

def save_to_xml(data, output_path):
    if not file_lock_wait(output_path):
        return False

    print_message(f'Now creating {os.path.basename(output_path)}')
    try:
        xml_bytes = dicttoxml.dicttoxml(data, custom_root='root', attr_type=False)
        xml_string = xml_bytes.decode('utf-8')
        with open(output_path, 'w', encoding='utf-8') as xml_file:
            xml_file.write(xml_string)
        print_message(f'Output saved to XML: {output_path}')
        return True
    except Exception as e:
        print_message(f"Error saving to XML: {e}")
        raise

def save_to_csv(data, output_path):
    if not file_lock_wait(output_path):
        return False

    print_message(f'Now creating {os.path.basename(output_path)}')
    flattened_data = []
    num_pages = len(data)
    page_num = 1
    for page in tqdm(data, f'Processing CSV row {page_num}...', num_pages):
    #for page in data:
        page_num += 1
        flat_page = {}
        for section, values in page.items():
            if isinstance(values, dict):
                for key, value in values.items():
                    if isinstance(value, dict):
                        for sub_key, sub_value in value.items():
                            if sub_key == "Description" and sub_value == key:
                                continue
                            if section != "Earnings" or sub_key != "Dates":
                                flat_page[f"{section}:{key}:{sub_key}"] = sub_value
                    else:
                        if key == "Description" and value == section:
                            continue
                        flat_page[f"{section}:{key}"] = value
            else:
                flat_page[section] = values
        flattened_data.append(flat_page)
    df = pd.DataFrame(flattened_data)
    try:
        df.to_csv(output_path, index=False)
        print_message(f"Output saved to CSV: {output_path}")
        return True
    except Exception as e:
        print_message(f"Error saving to CSV: {e}")
        raise

def flatten_dict(d, parent_key='', sep=':'):
    items = []
    if isinstance(d, list):
        for i, item in enumerate(d):
            items.extend(flatten_dict(item, f"{parent_key}{sep}{i}" if parent_key else str(i), sep=sep).items())
    elif isinstance(d, dict):
        for k, v in d.items():
            new_key = parent_key + sep + k if parent_key else k
            if isinstance(v, (dict, list)):
                items.extend(flatten_dict(v, new_key, sep=sep).items())
            else:
                items.append((new_key, v))
    return dict(items)

def create_hierarchical_headers(ws, df, fills, fonts):
    column_hierarchies = [col.split(':') for col in df.columns]
    max_levels = max(len(hierarchy) for hierarchy in column_hierarchies)

    for level in range(max_levels):
        row = []
        for hierarchy in column_hierarchies:
            if level < len(hierarchy):
                row.append(hierarchy[level])
            else:
                row.append('')
        ws.append(row)
        
        for col_idx, value in enumerate(row):
            cell = ws.cell(row=level + 1, column=col_idx + 1)
            cell.fill = fills[level % len(fills)]
            cell.font = fonts[level % len(fonts)]
            if level < max_levels - 1:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='left')

    # Apply gradient fill to the top row
    top_row_gradient_fill = GradientFill(degree=90, stop=("6376C7", "FFFFFF"))
    for col_idx in range(1, len(column_hierarchies) + 1):
        ws.cell(row=1, column=col_idx).fill = top_row_gradient_fill

    # Merge cells with identical entries in each header row
    for level in range(max_levels):
        row_values = [ws.cell(row=level+1, column=col_idx+1).value for col_idx in range(len(column_hierarchies))]
        start_idx = 0
        while start_idx < len(row_values):
            end_idx = start_idx
            while end_idx < len(row_values) and row_values[end_idx] == row_values[start_idx]:
                end_idx += 1
            if end_idx - start_idx > 1:
                ws.merge_cells(start_row=level+1, start_column=start_idx+1, end_row=level+1, end_column=end_idx)
                cell = ws.cell(row=level+1, column=start_idx+1)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            start_idx = end_idx

    # Merge cells with identical entries vertically across header rows
    for col_idx in range(len(column_hierarchies)):
        for level in range(max_levels):
            start_row = level + 1
            end_row = max_levels
            value = ws.cell(row=start_row, column=col_idx + 1).value
            if value:
                for row in range(start_row + 1, end_row + 1):
                    if ws.cell(row=row, column=col_idx + 1).value == value:
                        ws.cell(row=row, column=col_idx + 1).value = None
                        ws.merge_cells(start_row=start_row, start_column=col_idx + 1, end_row=row, end_column=col_idx + 1)
                    else:
                        break
    
    return max_levels
