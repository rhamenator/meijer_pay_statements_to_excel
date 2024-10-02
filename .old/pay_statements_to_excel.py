#! python3

#pay_statements_to_excel.py
import subprocess
import re
import sys
import os
import platform
import time
import pandas as pd
import numpy as np
import json
import dicttoxml # type: ignore
import fitz  # PyMuPDF
import keyboard
import tkinter as tk
#import tkinterdnd2
import traceback
from tkinter import Tk, filedialog, Text, Scrollbar, Button, END, Toplevel, Label, ttk, messagebox, IntVar
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Protection, PatternFill, Font, Alignment, NamedStyle, GradientFill
from openpyxl.worksheet.table import Table, TableStyleInfo
#from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm 

# Set up logging
# logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

class TqdmToText(tqdm):
    def __init__(self, *args, **kwargs):
        self.text_widget = kwargs.pop('text_widget', None)
        super().__init__(*args, **kwargs)

    def display(self, msg=None, pos=None):
        if self.text_widget:
            self.text_widget.insert(tk.END, self.format_meter(self.n, self.total, time.time() - self.start_t) + '\n')
            self.text_widget.see(tk.END)
        else:
            super().display(msg, pos)

class PrintLogger:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, message):
        self.text_widget.insert(tk.END, message)
        self.text_widget.see(tk.END)

    def flush(self):
        pass

# ErrorLogger class with additional method for showing error dialog
class ErrorLogger(PrintLogger):
    def showerror(self, message):
        messagebox.showerror("Error", message)

def os_supports_gui():
    os_name = platform.system()
    return os_name in ['nt', 'Windows', 'Darwin', 'macOS', 'Linux', 'FreeBSD', 'OpenBSD', 'NetBSD', 'SunOS']

def command_line_mode():
    if sys.stdin.isatty():
        return True
    if not os.environ.get('DISPLAY'):
        return True
    return False

def pick_input_file_gui():
    if command_line_mode():
        root = tk.Tk()
        root.withdraw()
    file_path = filedialog.askopenfilename(defaultextension=".pdf", title="Select the PDF file containing your pay statements", filetypes=[("PDF Files", "*.pdf")])
    if command_line_mode():
        if root is not None: root.destroy()
    return file_path

def pick_input_file_cli():
    file_path = input("Enter the path to the PDF file containing your pay statements: ")
    return file_path

def pick_output_file_gui(input_file):
    if command_line_mode():
        root = tk.Tk()
        root.withdraw()
    file_path = filedialog.asksaveasfilename(confirmoverwrite=True, initialfile=os.path.basename(input_file).replace('.pdf', '.xlsx'), defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Select the name of the Excel you want to create")
    if command_line_mode():
        if root is not None: root.destroy()
    return file_path

def pick_output_file_cli():
    file_path = input("Enter the name of the Excel workbook you want to create, or press [Enter] to cancel: ")
    return file_path

def pick_input_file(file_path):
    # global gui_mode
    while True:
        if file_path is None:
            file_path = pick_input_file_gui() if gui_mode or os_supports_gui() else pick_input_file_cli()
        if not file_path: return None
        if os.path.exists(file_path):
            if not gui_mode: print_message(f"File selected: {file_path}")
            return file_path
        else:
            if gui_mode or os_supports_gui():
                if command_line_mode():
                    root = tk.Tk()
                    root.withdraw()
                answer = messagebox.askretrycancel('File selected does not exist. Click Retry to try again, or Cancel to quit')
                if command_line_mode() and root is not None: root.destroy()
                if answer == 'retry':
                    file_path = ''
                    continue
                else:
                    return None
            else:
                print_message("File does not exist. Press any key to try again, or press [Enter], [ctrl/command]+[q], or [Esc] to quit.", user_input=True)
                event = keyboard.read_event()
                if event.event_type == keyboard.KEY_DOWN and event.name in ('ctrl+q', 'esc', 'enter'):
                    return None
                else:
                    file_path = ''
                    continue
                
def pick_output_file(input_file, file_path):
    # global gui_mode    
    while True:
        if file_path is None:
            if command_line_mode():
                root = tk.Tk()
                root.withdraw()
            file_path = pick_output_file_gui(input_file) if gui_mode or os_supports_gui() else pick_output_file_cli()
            if command_line_mode() and root is not None: root.destroy()
        if not file_path:
            return None
        if not os.path.exists(file_path):
            return file_path
        else:
            if gui_mode or os_supports_gui():
                return file_path
            else:
                overwrite = input(f"File {file_path} already exists. Do you want to overwrite it? (Press [y]='yes' or [n]=no, default=[n]): ").lower()
                if overwrite == 'y':
                    return file_path
                else:
                    file_path = None
                    continue

def print_message(message, title=None, user_input=False, dialog_type=None, duration=3000):
    # global gui_mode    
    if command_line_mode():
        root = tk.Tk()
        root.withdraw()
    response = None
    if user_input:
        if gui_mode or os_supports_gui():
            if dialog_type == "info" or dialog_type == None:
                messagebox.showinfo(title if title is not None else "Information", message)
            elif dialog_type == "showwarning":
                messagebox.showwarning(title if title is not None else "Warning", message)
            elif dialog_type == "showerror":
                messagebox.showerror(title if title is not None else "Error", message)
            elif dialog_type == "askokcancel":
                response = messagebox.askokcancel(title if title is not None else "Question", message)
            elif dialog_type == "askyesno":
                response = messagebox.askyesno(title if title is not None else "Question", message)
            elif dialog_type == "askyesnocancel":
                response = messagebox.askyesnocancel(title if title is not None else "Question", message)
            elif dialog_type == "askretrycancel":
                response = messagebox.askretrycancel(title if title is not None else "Problem", message)
            elif user_input:
                if root is None and gui_mode: 
                    show_temporary_message(message, duration=3000)
                elif root is not None and gui_mode:
                    show_in_main_window(title, message)
                else: print(message)
            else: show_in_main_window(title, message)
        else:
            print(message)
    else:
        if root is None and gui_mode: 
            show_temporary_message(message, duration=3000)
        elif root is not None and gui_mode:
            show_in_main_window(title, message)
        else: print(message)
    if command_line_mode() and root is not None: root.destroy()
    return response

def show_in_main_window(title, message):
    # Update or create a label in Tk.root to display the message
    # global root
    if root is None:
        # print("Error: Tk.root is not initialized.")
        return
    # Clear any existing message label
    clear_main_window_message()
    
    # Create a label to display the message
    label_message = tk.Label(root, text=message, wraplength=300)
    label_message.pack(pady=20)
    
    # Optionally set a title for the main window
    root.title(title if title else "Messages:")

def clear_main_window_message():
    # Function to clear the message label in Tk.root
    # global root
    for widget in root.winfo_children():
        widget.destroy()

def show_temporary_message(message, duration=3000):
    top = Toplevel()
    top.title("Message")
    top.geometry("300x100+100+100")
    Label(top, text=message, wraplength=280).pack(expand=True)
    top.after(duration, top.destroy)

def is_file_locked(filepath):
    if not os.path.exists(filepath):
        return False
    try:
        with open(filepath, 'a'):
            pass
    except IOError:
        return True
    return False

# Function to analyze text on a page and fill the data dictionary
def analyze_page_text(page_text, page_num): #, total_pages):
    #global gui_mode, progress_bar, progress
    data = {
        "Identification": {},
        "Pay Period Info": {},
        "Earnings": {},
        "Employee Taxes": {},
        "Post Tax Deductions": {},
        "Pre Tax Deductions": {},
        "Employer Paid Benefits": {},
        "Subject or Taxable Wages": {},
        "Allowances": {},
        "Absence Plans": {},
        "Payment Information": {}
    }
    
    lines = []
    section_names = ["Earnings", "Employee Taxes", "Pre Tax Deductions", "Post Tax Deductions", "Subject or Taxable Wages", "Employer Paid Benefits", "Absence Plans", "Marital Status", "Payment Information"]

    for block in page_text:
        lines.append(block[4])

    # Process Pay Period Info
    if len(lines) > 4:
        values = lines[3].split('\n')
        identification_headers = ["Name", "Company", "Employee ID"]
        identification_values = values[0:2]
        data["Identification"].update(dict(zip(identification_headers, identification_values)))
        pay_period_info_headers = ["Pay Period Begin", "Pay Period End", "Check Date", "Check Number"]
        pay_period_info_values = values[3:6]
        data["Pay Period Info"].update(dict(zip(pay_period_info_headers, pay_period_info_values)))
        pay_period_begin = values[3]
        pay_period_end = values[4]

    # Process Current and YTD Summary
    if len(lines) > 6:
        # Split summary data
        summary_field_names = lines[4].split('\n')
        summary_current_values = lines[5].split('\n')[1:]  # Exclude the first entry which is "Current"
        summary_ytd_values = lines[6].split('\n')[1:]  # Exclude the first entry which is "YTD"
        
        # Ensure lengths match
        if len(summary_field_names) != len(summary_current_values) or len(summary_field_names) != len(summary_ytd_values):
            raise ValueError("Summary data lengths do not match.")

        # Add data to "Current" and "YTD" sub-dictionaries
        data["Pay Period Info"]["Current"] = {}
        data["Pay Period Info"]["YTD"] = {}
        
        for field, current, ytd in zip(summary_field_names, summary_current_values, summary_ytd_values):
            field = field.strip()
            if field:
                current_data = current.replace(",", "").strip()
                ytd_data = ytd.replace(",", "").strip()
                data["Pay Period Info"]["Current"][field] = safe_float_conversion(current_data)
                data["Pay Period Info"]["YTD"][field] = safe_float_conversion(ytd_data)
           
    # Other sections
    if len(lines) > 7:
        line_no = 7
        for line in lines[line_no:]:
            this_line = lines.index(line)
            if line == "Earnings\n":
                other_sections = section_names
                other_sections.remove(line.strip("\n"))
                # print(f"Processing {line} section of page {page_num + 1}, line {this_line}")
                process_earnings_table(lines[this_line:], data["Earnings"], page_num + 1, pay_period_begin, pay_period_end, other_sections)
            elif line == "Employee Taxes\n":
                other_sections = section_names
                other_sections.remove(line.strip("\n"))
                # print(f"Processing {line} section of page {page_num + 1}, line {this_line}")
                process_deductions(lines[this_line:], data["Employee Taxes"], page_num + 1, other_sections)
            elif line == "Pre Tax Deductions\n":
                other_sections = section_names
                other_sections.remove(line.strip("\n"))
                # print(f"Processing {line} section of page {page_num + 1}, line {this_line}")
                process_deductions(lines[this_line:], data["Pre Tax Deductions"], page_num + 1, other_sections)
            elif line == "Post Tax Deductions\n":
                other_sections = section_names
                other_sections.remove(line.strip("\n"))
                # print(f"Processing {line} section of page {page_num + 1}, line {this_line}")
                process_deductions(lines[this_line:], data["Post Tax Deductions"], page_num + 1, other_sections)
            elif line == "Subject or Taxable Wages\n":
                other_sections = section_names
                other_sections.remove(line.strip("\n"))
                # print(f"Processing {line} section of page {page_num + 1}, line {this_line}")
                process_table_subject_taxable_wages(lines[this_line:], data["Subject or Taxable Wages"], page_num + 1, other_sections)
            elif line == "Employer Paid Benefits\n":
                other_sections = section_names
                other_sections.remove(line.strip("\n"))
                # print(f"Processing {line} section of page {page_num + 1}, line {this_line}")
                process_employer_paid_benefits(lines[this_line:], data["Employer Paid Benefits"], page_num + 1, other_sections)
            elif "Absence Plans" in line:
                other_sections = section_names
                other_sections.remove("Absence Plans")
                # print(f"Processing {line} section of page {page_num + 1}, line {this_line}")
                process_absence_plans(lines[this_line:], data["Absence Plans"], page_num + 1, other_sections)
            elif "Marital Status" in line:
                other_sections = section_names
                other_sections.remove("Marital Status")
                # print(f"Processing Allowances section of page {page_num + 1}, line {this_line}")
                process_allowances(lines[this_line:], data["Allowances"], page_num + 1, other_sections)
            elif line == "Payment Information\n":
                other_sections = section_names
                other_sections.remove(line.strip("\n"))
                # print(f"Processing {line} section of page {page_num + 1}, line {this_line}")
                process_payment_information(lines[this_line:], data["Payment Information"], page_num + 1, other_sections)
    return data

def safe_float_conversion(value):
    try:
        if type(value) == float:
            return value
        if type(value) == str:
            result = 0.0 if value.isspace() else float(value)
            return result
        if type(value) == bool:
            result = 1 if value == True else 0
            return result
        return float(value)
    except ValueError:
        return 0.0

def safe_int_conversion(value):
    try:
        if type(value) == int: 
            return value
        if type(value) == float:
            return int(value)
        if type(value) == str:
            result = 0 if value.isspace() else int(value)
            return result
        if type(value) == bool:
            result = 1 if value == True else 0
            return result
        return int(value)
    except ValueError:
        return 0

def process_earnings_table(lines, target_dict, page_num, pay_period_begin, pay_period_end, section_names):
    dictionary_name = ''
    ending_tags = ['Total']
    ending_tags.extend(section_names)
    for line_no, line in enumerate(lines):
        values = line.split('\n')
        if line_no in [0,1]:
            continue
        elif values[0] in ending_tags:
            break
        first_col_updated = ''
        dates = ''
        hours = ''
        rate = 0
        amount = 0
        ytd_hours = 0
        ytd_amount = 0
        if line_no == 0:
            dictionary_name = line.strip('\n')
        if len(values) >= 2:
            description = values[0]
            date_regex = r'^(.*?)\b(\d{2}/\d{2}/\d{4})\b'
            match = re.match(date_regex, description)
            if "Total" in values[0]:
                description = (((values[0]).strip()).replace(':','')).replace(dictionary_name,'')
                dates = pay_period_begin + " - " + pay_period_end
                amount = safe_float_conversion(values[1].replace(",", ""))
                ytd_amount = safe_float_conversion(values[2].replace(",", ""))
                target_dict[description] = {'Dates': dates, 'Amount': amount, 'YTD Amount': ytd_amount}
                continue
            elif match:
                dates = description[-23:]
                description_text = description.replace(dates,"")
                first_col_updated = description_text.strip()
                if len(values) < 8:
                    values.insert(1,dates)
                else:
                    values[1]=dates
                if values[2] == "0" and (values[3]).isspace() and (values[4]).isspace():
                    continue
                else: hours = safe_float_conversion(values[2])
                rate = safe_float_conversion(values[3])
                amount_text = values[4].replace(",", "")
                if amount_text.count(".") > 1:
                    amount = safe_float_conversion(amount_text[:amount_text.find(".") + 3])
                    ytd_hours = safe_float_conversion(amount_text.replace((values[4].replace(",", "")).replace(amount_text,""),""))
                    ytd_amount = safe_float_conversion(values[5].replace(",", ""))                
                else:
                    amount = safe_float_conversion(amount_text)
                    ytd_hours = safe_float_conversion(values[5].replace(",", ""))
                    ytd_amount = safe_float_conversion(values[6].replace(",", ""))
            else:
                first_col_updated = description
                if len(values) < 8 and not values[1].strip():
                    values.insert(1,pay_period_begin + " - " + pay_period_end)
                dates = (values[1]).strip()
                
                if len((values[2]).split()) > 1:
                    split_values = (values[2]).split()
                    hours = safe_float_conversion(split_values[0])
                    values.insert(3, split_values[1])
                else:
                    hours = safe_float_conversion(values[2])
                rate = safe_float_conversion(values[3])
                amount_text = values[4].replace(",", "")
                if amount_text.count(".") > 1:
                    amount = safe_float_conversion(amount_text[:amount_text.find(".") + 3])
                    ytd_hours_text = amount_text.replace((values[4].replace(",", "")).replace(amount_text,""),"")
                    if " " in ytd_hours_text:
                        ytd_hours = safe_float_conversion(ytd_hours_text[ytd_hours_text.find(" ")+1:])
                    else:
                        ytd_hours = safe_float_conversion((amount_text[amount_text.find(".") + 3:]).strip())
                    ytd_amount = safe_float_conversion(values[5].replace(",", ""))                
                else:
                    amount = safe_float_conversion(amount_text)
                    if first_col_updated == "PRC Hours Balance":
                        ytd_hours = 0
                        ytd_amount = safe_float_conversion(values[5].replace(",", ""))
                    else:
                        if len(values) < 8:
                            ytd_hours = safe_float_conversion(values[4].replace(",", ""))
                            ytd_amount = safe_float_conversion(values[5].replace(",", ""))
                        else: 
                            ytd_hours = safe_float_conversion(values[5].replace(",", ""))
                            ytd_amount = safe_float_conversion(values[6].replace(",", ""))

            if first_col_updated == "Night Premium" and rate:
                first_col_updated = "Night Premium"

            if first_col_updated in target_dict:
                existing_data = target_dict[first_col_updated]
                existing_data['Hours'] += hours
                existing_data['Amount'] += amount
                existing_data['YTD Hours'] += ytd_hours
                existing_data['YTD Amount'] += ytd_amount
            else:
                target_dict[first_col_updated] = {'Dates': dates, 'Hours': hours, 'Rate': rate, 'Amount': amount, 'YTD Hours': ytd_hours, 'YTD Amount': ytd_amount}

            if not dates.strip():
                target_dict[first_col_updated]['Amount'] = 0
        
        else: break
        
def process_deductions(lines, target_dict, page_num, other_sections):
    headers = ''
    dictionary_name = ''
    ending_tags = ['Total']
    ending_tags.extend(other_sections)
    for line_no, line in enumerate(lines):
        if line_no == 0:
            dictionary_name = line.strip('\n')
        elif line_no == 1:
            headers = line.split('\n')
            continue
        else:
            values = line.split('\n')
            description = values[0].strip()
            if description.endswith('Total:'):
                description = 'Total'
            if len(values) > 2:
                values[1] = safe_float_conversion(values[1].replace(",", ""))
                values[2] = safe_float_conversion(values[2].replace(",", ""))
                section_data = dict(zip(headers[1:-1], values[1:-1]))
                target_dict[description] = {k: v for k, v in section_data.items()}
            if description == 'Total' or description in ending_tags: break

def process_employer_paid_benefits(lines, target_dict, page_num, other_sections):
    headers = ''
    dictionary_name = ''
    ending_tags = ['Total']
    ending_tags.extend(other_sections)
    for line_no, line in enumerate(lines):
        if line_no == 0:
            dictionary_name = line.strip('\n')
        elif line_no == 1:
            headers = line.split('\n')
            continue
        else:
            values = line.split('\n')
            description = values[0]
            if len(values) > 2:
                values[1] = safe_float_conversion(values[1].replace(",", ""))
                values[2] = safe_float_conversion(values[2].replace(",", ""))
                section_data = dict(zip(headers[1:-1], values[1:-1]))
                if 'Total' in description: description = 'Total'
                target_dict[description] = {k: v for k, v in section_data.items()}
            if description in ending_tags: break

def process_table_subject_taxable_wages(lines, target_dict, page_num, other_sections):
    headers = ''
    dictionary_name = ''
    ending_tags = ['Total', 'Federal']
    ending_tags.extend(other_sections)
    for line_no, line in enumerate(lines):
        values = line.split('\n')
        if ((values[0]).replace('\n','')).strip() in ending_tags:
            break
        elif line_no == 0:
            dictionary_name = values[0]
            continue
        elif line_no == 1:
            headers = values
            continue
        else:
            for item in range(1, len(values[1:])):
                values[item] = safe_float_conversion(values[item].replace(',',''))
            section_data = dict(zip(headers[1:-1], values[1:-1]))
            description = values[0]
            target_dict[description] = {k: v for k, v in section_data.items()}

def process_allowances(lines, target_dict, page_num, other_sections):
    ending_tags = ['Total', 'Payment Information']
    ending_tags.extend(other_sections)
    for line_no, line in enumerate(lines):
        values = line.split('\n')
        if line_no == 0:
            headers = ['Description', 'Federal', 'State']
        description = (values[0]).strip()
        if description == ending_tags[1]:
            break
        for item in range(1, len(values[1:])):
            values[item] = safe_int_conversion(values[item].replace(',',''))
        section_data = dict(zip(headers[1:], values[1:]))
        target_dict[description] = {k: v for k, v in section_data.items()}
        if description in ending_tags: break

def process_absence_plans(lines, target_dict, page_num, other_sections):
    dictionary_name = ''
    ending_tags = ['Total', 'Federal']
    ending_tags.extend(other_sections)
    for line_no, line in enumerate(lines):
        this_line = line.split('\n')
        if line_no == 0:
            dictionary_name = this_line[0]
            continue
        elif line_no == 1:
            headers = this_line
            continue
        elif this_line[0] in ending_tags:
            break
        else:
            values = line.split('\n')
            for item in range(1, len(values[1:])):
                values[item] = safe_float_conversion(values[item].replace(',',''))
            section_data = dict(zip(headers[1:-1], values[1:-1]))
            description = headers[0]
            target_dict[description] = {k: v for k, v in section_data.items()}

def process_payment_information(lines, target_dict, page_num, other_sections):
    ending_tags = ['Total', 'Federal']
    ending_tags.extend(other_sections)
    headers = []
    values=[]
    description = ''
    headers = ''
    dictionary_name = ''
    for line_no, line in enumerate(lines):
        this_line = line.split('\n')
        if line_no == 0:
            dictionary_name = this_line[0]
            continue
        elif line_no == 1:
            headers = this_line
        else:
            values = line.split('\n')
            description = values[0]
            if description in ending_tags: break
            else:
                headers[0] = 'Name'
                if (values[-2]).endswith('USD'):
                    amountfield = (values[-2]).split()
                    values[-2] = safe_float_conversion(amountfield[0])
                    values[-1] = safe_float_conversion(amountfield[0])
                    values.append(amountfield[1])
                    headers[-1] = 'Currency'
                else:
                    amountfield = (values[-1]).split()
                    values[-1] = amountfield[0]
                    values.append(amountfield[1])
                    headers[-1] = 'Currency'
                section_data = dict(zip(headers, values))
                target_dict['Bank'] = {k: v for k, v in section_data.items()}
        if description in ending_tags: break

def save_to_json(data, output_path):
    while is_file_locked(output_path):
        if gui_mode:
            err_message = f'{os.path.basename(output_path)} is locked by another application. Please close it and retry or cancel'
            response = messagebox.askretrycancel(title='Error', message=err_message)
            if response == 'cancel' or response == None:
                return False
        else:
            print_message(f'{os.path.basename(output_path)} is locked by another application. Please close it then press a key, or press "ctrl+q" or "esc" to quit...', user_input=True)
            event = keyboard.read_event()
            if event.event_type == keyboard.KEY_DOWN and event.name in ('ctrl+q', 'esc'):
                return False
            time.sleep(1)

    print_message(f'Now creating {os.path.basename(output_path)}')
    try:
        with open(output_path, 'w', encoding='utf-8') as json_file:
            json.dump(data, json_file, ensure_ascii=False, indent=4)
        print_message(f'Output saved to JSON: {os.path.basename(output_path)}')
        return True
    except Exception as e:
        print_message(f"Error saving to JSON: {e}")
        raise

def save_to_xml(data, output_path):
    while is_file_locked(output_path):
        if gui_mode:
            err_message = f'{os.path.basename(output_path)} is locked by another application. Please close it and retry or cancel'
            response = messagebox.askretrycancel(title='Error', message=err_message)
            if response == 'cancel' or response == None:
                return False
        else:
            print_message(f'{os.path.basename(output_path)} is locked by another application. Please close it then press a key, or press "ctrl+q" or "esc" to quit...', user_input=True)
            event = keyboard.read_event()
            if event.event_type == keyboard.KEY_DOWN and event.name in ('ctrl+q', 'esc'):
                return False
            time.sleep(1)

    print_message(f'Now creating {os.path.basename(output_path)}')
    try:
        xml_bytes = dicttoxml.dicttoxml(data, custom_root='root', attr_type=False)
        xml_string = xml_bytes.decode('utf-8')
        with open(output_path, 'w', encoding='utf-8') as xml_file:
            xml_file.write(xml_string)
        print_message(f'Output saved to XML: {output_path}')
        return True
    except Exception as e:
        print_message(f"Error saving to XML: {e}")
        raise

def save_to_csv(data, output_path):
    # global gui_mode
    while is_file_locked(output_path):
        if gui_mode:
            err_message = f'{os.path.basename(output_path)} is locked by another application. Please close it and retry or cancel'
            response = messagebox.askretrycancel(title='Error', message=err_message)
            if response == 'cancel' or response == None:
                return False
        else:
            print_message(f'{os.path.basename(output_path)} is locked by another application. Please close it then press a key, or press "ctrl+q" or "esc" to quit...', user_input=True)
            event = keyboard.read_event()
            if event.event_type == keyboard.KEY_DOWN and event.name in ('ctrl+q', 'esc'):
                return False
            time.sleep(1)

    print_message(f'Now creating {os.path.basename(output_path)}')
    flattened_data = []
    for page in data:
        flat_page = {}
        for section, values in page.items():
            if isinstance(values, dict):
                for key, value in values.items():
                    if isinstance(value, dict):
                        for sub_key, sub_value in value.items():
                            if sub_key == "Description" and sub_value == key:
                                continue
                            if section != "Earnings" or sub_key != "Dates":
                                flat_page[f"{section}:{key}:{sub_key}"] = sub_value
                    else:
                        if key == "Description" and value == section:
                            continue
                        flat_page[f"{section}:{key}"] = value
            else:
                flat_page[section] = values
        flattened_data.append(flat_page)
    df = pd.DataFrame(flattened_data)
    try:
        df.to_csv(output_path, index=False)
        print_message(f"Output saved to CSV: {output_path}")
        return True
    except Exception as e:
        print_message(f"Error saving to CSV: {e}")
        raise

def flatten_dict(d, parent_key='', sep=':'):
    items = []
    for k, v in d.items():
        new_key = parent_key + sep + k if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        else:
            items.append((new_key, v))
    return dict(items)

def create_hierarchical_headers(ws, df, fills, fonts):
    column_hierarchies = [col.split(':') for col in df.columns]
    max_levels = max(len(hierarchy) for hierarchy in column_hierarchies)

    for level in range(max_levels):
        row = []
        for hierarchy in column_hierarchies:
            if level < len(hierarchy):
                row.append(hierarchy[level])
            else:
                row.append('')
        ws.append(row)
        
        for col_idx, value in enumerate(row):
            cell = ws.cell(row=level + 1, column=col_idx + 1)
            cell.fill = fills[level % len(fills)]
            cell.font = fonts[level % len(fonts)]
            if level < max_levels - 1:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='left')

    # Apply gradient fill to the top row
    top_row_gradient_fill = GradientFill(degree=90, stop=("6376C7", "FFFFFF"))
    for col_idx in range(1, len(column_hierarchies) + 1):
        ws.cell(row=1, column=col_idx).fill = top_row_gradient_fill

    # Merge cells with identical entries in each header row
    for level in range(max_levels):
        row_values = [ws.cell(row=level+1, column=col_idx+1).value for col_idx in range(len(column_hierarchies))]
        start_idx = 0
        while start_idx < len(row_values):
            end_idx = start_idx
            while end_idx < len(row_values) and row_values[end_idx] == row_values[start_idx]:
                end_idx += 1
            if end_idx - start_idx > 1:
                ws.merge_cells(start_row=level+1, start_column=start_idx+1, end_row=level+1, end_column=end_idx)
                cell = ws.cell(row=level+1, column=start_idx+1)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            start_idx = end_idx

    # Merge cells with identical entries vertically across header rows
    for col_idx in range(len(column_hierarchies)):
        for level in range(max_levels):
            start_row = level + 1
            end_row = max_levels
            value = ws.cell(row=start_row, column=col_idx + 1).value
            if value:
                for row in range(start_row + 1, end_row + 1):
                    if ws.cell(row=row, column=col_idx + 1).value == value:
                        ws.cell(row=row, column=col_idx + 1).value = None
                        ws.merge_cells(start_row=start_row, start_column=col_idx + 1, end_row=row, end_column=col_idx + 1)
                    else:
                        break
    
    return max_levels

def save_to_excel(data, output_path):
    # global gui_mode, progress_bar, progress
    while is_file_locked(output_path):
        if gui_mode:
            err_message = f'{os.path.basename(output_path)} is locked by another application. Please close it and retry or cancel'
            response = messagebox.askretrycancel(title='Error', message=err_message)
            if response == 'cancel' or response == None:
                return False
        else:
            print_message(f'{os.path.basename(output_path)} is locked by another application. Please close it then press a key, or press "ctrl+q" or "esc" to quit...', user_input=True)
            event = keyboard.read_event()
            if event.event_type == keyboard.KEY_DOWN and event.name in ('ctrl+q', 'esc'):
                return False
            time.sleep(1)

    print_message(f'Now creating {output_path}')
    num_pages = len(data)
    flattened_data = []
    for idx, page in enumerate(data):
        flat_page = flatten_dict(page)
        flat_page['UniqueID'] = idx + 1
        flattened_data.append(flat_page)

    df = pd.DataFrame(flattened_data)

    try:
        wb = Workbook()

        accounting_format = NamedStyle(name='accounting_format', number_format='_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)')
        date_format = NamedStyle(name='date_format', number_format='mm/dd/yyyy')
        number_format = NamedStyle(name='number_format', number_format='#,##0.00')
        wb.add_named_style(accounting_format)
        wb.add_named_style(date_format)
        wb.add_named_style(number_format)

        def is_numeric_string(s):
            return bool(re.match(r'^\d+$', s))

        top_level_keys = ["Identification", "Pay Period Info", "Earnings", "Employee Taxes", "Post Tax Deductions", "Pre Tax Deductions", "Employer Paid Benefits", "Subject or Taxable Wages", "Allowances", "Absence Plans", "Payment Information"]
        hair_border = Border(left=Side(style='hair'), 
                     right=Side(style='hair'), 
                     top=Side(style='hair'), 
                     bottom=Side(style='hair'))
        fills = [
            PatternFill(start_color="6376C7", end_color="6376C7", fill_type="solid"),  
            PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid"),  # Light Cyan
            PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Light Green
            PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")   # Light Yellow
        ]
        fonts = [
            Font(size=12, bold=True),
            Font(size=11, bold=True),
            Font(size=10, bold=True),
            Font(size=9, bold=True)
        ]

        for section in top_level_keys:
            if section == "General":
                continue

            section_df = df[[col for col in df.columns if col.startswith(section + ':') or col.endswith("Date") or col.endswith("Begin") or col.endswith("End")]]
            if section_df.empty:
                continue
            
            ws = wb.create_sheet(title=section)
            
            max_levels = create_hierarchical_headers(ws, section_df, fills, fonts)
            # row_num = 0
            outfile = os.path.basename(output_path)
            rows = dataframe_to_rows(section_df, index=False, header=False)
            for r in tqdm(rows, f'Appending data to the {section} sheet of {outfile}', rows.__sizeof__()):
                # row_num += 1
                # print_message(f'Appending data to the {section} sheet of {output_path}, row: {row_num}', user_input=False)
                ws.append(r)
            total_rows = ws.max_row
            # row_num = 0
            # if gui_mode:
            #     update_progress_bar(row_num, total_rows)
            # print_message(f'Formatting cells in the {section} sheet of {output_path}, {total_rows} rows...', user_input=False)
            sheet_rows = ws.iter_rows(min_row=max_levels + 1, max_row=total_rows, min_col=1, max_col=ws.max_column)
            for row in tqdm(sheet_rows, f'Formatting cells in the {section}', total_rows):
                # row_num += 1
                # if gui_mode:
                #     update_progress_bar(row_num, total_rows)
                for cell in row:
                    col_name = section_df.columns[cell.column - 1]
                    if cell.row <= max_levels:
                        cell.alignment = Alignment(horizontal='center')
                    else:
                        if col_name.endswith(('Rate', 'Amount', 'YTD', 'Pay', 'Deductions', 'Taxes')):
                            cell.style = accounting_format
                        elif 'date' in col_name.lower() or 'begin' in col_name.lower() or 'start' in col_name.lower() or 'end' in col_name.lower():
                            cell.style = date_format
                        elif isinstance(cell.value, float):
                            cell.style = number_format
                        elif isinstance(cell.value, int):
                            cell.number_format = '0'
                        elif isinstance(cell.value, str) and is_numeric_string(cell.value):
                            cell.number_format = '0'
            
            # print_message(f'Adjusting column widths in {outfile}', user_input=False)
            # num_columns = ws.max_column
            # column_number = 0
            # if gui_mode:
            #     update_progress_bar(column_number, num_columns)
            columns = ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column)
            for col in tqdm(columns,f'Adjusting column widths in the {section}', ws.max_column):
                # column_number += 1
                # if gui_mode:
                #     update_progress_bar(column_number, num_columns)
                max_length = max(len(str(cell.value)) for cell in col)
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width
            
            dark_row_fill_color = '98E3E2'
            light_row_fill_color = 'C5E3E2'
            light_shade = PatternFill(start_color=light_row_fill_color, end_color=light_row_fill_color, fill_type="solid")
            medium_shade = PatternFill(start_color=dark_row_fill_color, end_color=dark_row_fill_color, fill_type="solid")
            # print_message(f'Formatting rows of {outfile}', user_input=False)
            # row_num = 0
            # if gui_mode:
            #     update_progress_bar(row_num, total_rows)
            for row in tqdm(range(max_levels + 1, ws.max_row + 1), f'Formatting rows of {outfile}', ws.max_row + 1):
                # row_num += 1
                # if gui_mode:
                #     update_progress_bar(row_num, total_rows)
                fill = light_shade if (row - max_levels) % 2 == 0 else medium_shade
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill

        wb.remove(wb['Sheet'])

        while is_file_locked(output_path):
            if gui_mode:
                err_message = f'{os.path.basename(output_path)} is locked by another application. Please close it and retry or cancel'
                response = messagebox.askretrycancel(title='Error', message=err_message)
                if response == 'cancel' or response == None:
                    return False
            else:
                print_message(f'{os.path.basename(output_path)} is locked by another application. Please close it then press a key, or press "ctrl+q" or "esc" to quit...', user_input=True)
                event = keyboard.read_event()
                if event.event_type == keyboard.KEY_DOWN and event.name in ('ctrl+q', 'esc'):
                    return False
                time.sleep(1)
        
        print_message(f'Now saving {os.path.basename(output_path)}')
        wb.save(output_path)
        print_message(f"{os.path.basename(output_path)} has been created.")
                    
    except Exception as e:
        print_message(f"Error creating {os.path.basename(output_path)}: {e}")
        raise

def extract_structured_data(file_path, text_widget = None):
    document = fitz.open(file_path)
    total_pages = len(document)
    output_data = []
    infile = os.path.basename(file_path)
    page_num = 0
    for page in tqdm(document,f'Reading and analyzing file {infile}', total_pages):
        page_num += 1
        page_text = page.get_text("blocks")
        output_data.append(analyze_page_text(page_text, page_num))
    return output_data
 
def main_logic(input_file, output_file, text_box = None):
    # global gui_mode
    print_message(f"Analyzing...\nInput file: {input_file}\nOutput file: {output_file}")
    file_path = input_file
    output_data = extract_structured_data(file_path, text_box)
    
    print_message(f"Creating JSON file from\nInput file: {input_file}")
    json_output_path = output_file.replace('.xlsx', '.json')
    save_to_json(output_data, json_output_path)
    
    print_message(f"Creating XML file from\nInput file: {input_file}")
    xml_output_path = output_file.replace('.xlsx', '.xml')
    save_to_xml(output_data, xml_output_path)
    
    print_message(f"Creating CSV file from\nInput file: {input_file}")
    csv_output_path = output_file.replace('.xlsx', '.csv')
    save_to_csv(output_data, csv_output_path)
    
    print_message(f"Creating Excel workbook\n{output_file}")
    excel_output_path = output_file
    success = save_to_excel(output_data, excel_output_path)
    if gui_mode:
        messagebox.showinfo('Information', 'All finished!')
        print_message('Close window to exit.')
    # else:
    #     print('Press any key to exit...')
    #     event = keyboard.read_event()
    return success

def run(input_file, output_file, text_box=None):
    if input_file is None:
        input_file = pick_input_file(input_file)
    if not input_file:
        print_message("No input file selected.")
        return False
    if output_file is None:
        output_file = pick_output_file(input_file, output_file)
    if not output_file:
        print_message("No output file selected.")
        return False
    return main_logic(input_file, output_file, text_box)
    
def main(*args):
    global gui_mode, root
    if len(args) > 0:
        input_file = args[1] if len(args) > 1 else None
        output_file = args[2] if len(args) > 2 else None
    gui_mode = os_supports_gui() and not sys.stdin.isatty()
    cli_mode = command_line_mode()
    if cli_mode:
        try:
            root = None
            run(input_file, output_file)
        except Exception as e:
            print(f"An error occurred: {e}", file=sys.stderr)
            traceback.print_exc(file=sys.stderr)  # Print the traceback
            if root:
                ErrorLogger.showerror(f"An error occurred:\n{e}")
                if root is not None: root.destroy()
            raise
    elif gui_mode:
        root = tk.Tk()
        root.title("Export pay statements to Excel")
        root.geometry("640x300")
        text_box = Text(root, wrap='none', height=300, width=600)
        text_box.pack(side='left', fill='both', expand=True)
        scrollbar = Scrollbar(root, command=text_box.yview)
        scrollbar.pack(side='right', fill='y')
        text_box.config(yscrollcommand=scrollbar.set)
        sys.stdout = PrintLogger(text_box)
        error_logger = ErrorLogger(text_box)  # Use ErrorLogger for error handling
        try:
            run(input_file, output_file)
            root.mainloop()
            return
        except Exception as e:
            print(f"An error occurred: {e}", file=sys.stderr)
            traceback.print_exc(file=sys.stderr)
            if root:
                ErrorLogger.showerror(f"An error occurred:\n{e}")
                if root is not None: root.destroy()
            raise
    else:
        try:
            root = None
            run(input_file, output_file)
        except Exception as e:
            print(f"An error occurred: {e}", file=sys.stderr)
            traceback.print_exc(file=sys.stderr)
            if root:
                ErrorLogger.showerror(f"An error occurred:\n{e}")
                if root is not None: root.destroy()
            raise
       
if __name__ == "__main__":
    try:
        root = None
        gui_mode = None
        main(sys.argv)
        sys.exit(0)  # Exit with code 0 for success
    except Exception as e:
        print(f"An error occurred: {e}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        if root:
            if root is not None: root.destroy()  # Close the GUI if it's open
        sys.exit(1)  # Exit with code 1 for failure